import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill
import io
import zipfile
import warnings
warnings.filterwarnings('ignore')
from datetime import datetime

def app():
    # Streamlit app setup
    st.markdown("<h1 style='text-align: center; width: 100%;'>Form003004 File Uploader/Downloader</h1>", unsafe_allow_html=True)
    # Add instruction for refreshing the page
    st.markdown("<p style='text-align: left; color: blue; font-family: Arial;'>For refresh and remove history, refresh from Browser Left top most corner.</p>", unsafe_allow_html=True)

    # Function to browse and select a file
    def browse_file(file_type):
        file_path = st.file_uploader(f"Upload {file_type} File:", type=["xlsx", "xls"], key=f"{file_type.lower().replace(' ', '_')}")
        if file_path:
            st.success(f"{file_type} File uploaded successfully!")
            return file_path

    # Browse and upload files
    file_scdb_path = browse_file("SCDB")
    file_subcontractor_path = browse_file("Subcontractor")

    # This function formatting for the subcontractor report
    def apply_formatting(writer, sheet_name, dataframe):
        # Access the worksheet
        ws = writer.sheets[sheet_name]
 
        # Define the fill pattern for the red color
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
 
        # Iterate through each row in the DataFrame and check the condition
        for row_number, row in dataframe.iterrows():
            # Check if the condition is met in the 'Remark' column
            if any(keyword in row['Remark'] for keyword in ['No Match', 'Closed']):
                # Apply the red fill to the entire row
                for col_number, value in enumerate(row):
                    ws.cell(row=row_number + 2, column=col_number + 1).fill = red_fill
    # Execute logic and download
    if st.button("Execute Program"):

        if file_scdb_path and file_subcontractor_path:
            try:
                st.info("Executing Programme for files generation ...")
            
                # Load dataframes
                # Master file downloaded from SCDB
                selected_columns = ['Task ID', 'Task Model (Name)', 'Asset - Tag', 'Task Status']
                df_master = pd.read_excel(file_scdb_path, dtype=str, usecols=selected_columns)
 
                # Subcon Data
                df_subcon = pd.read_excel(file_subcontractor_path, dtype=str, skiprows=1)
                df_subcon.columns=['Report/RFI No','Parameter Reading 01', 'Parameter Reading 02', 'Parameter Reading 03', 'Parameter Reading 04', 'Parameter Reading 05', 'Parameter Reading 06', 'Parameter Reading 07', 'Parameter Reading 08', 'Parameter Reading 09', 'Parameter Reading 10', 'Completed Date', 'Completed By']
                # Add columns at the start of df_subcon , and the column names are ['Step Sequence', 'Step Point', 'Inspection Type']
                df_subcon.insert(0, 'Step Sequence', '')
                df_subcon.insert(1, 'Step Point', '')
                df_subcon.insert(2, 'Inspection Type', '')
                # Correcting the Date format based user requirment
                df_subcon[['Parameter Reading 04','Parameter Reading 07', 'Completed Date']] = df_subcon[['Parameter Reading 04','Parameter Reading 07', 'Completed Date']].apply(pd.to_datetime, errors='coerce')
                df_subcon[['Parameter Reading 04','Parameter Reading 07', 'Completed Date']] = df_subcon[['Parameter Reading 04','Parameter Reading 07', 'Completed Date']].apply(lambda x: x.dt.strftime('%m/%d/%Y'))
                df_master.columns=['Task ID', 'Task Model (Name)', 'Asset - Tag', 'Task Status']
                result_df = pd.merge(df_master, df_subcon,how='right', indicator=True, left_on='Asset - Tag', right_on='Parameter Reading 01')
 
                # not_matching_rows = result_df.loc[result_df['_merge'] == 'right_only', ['Task ID', 'Cable\nDrum\nNo.']]
                result_df['Remark'] = 'Good Match'
                # can be select a row from df where column vale is Nan
                result_df.loc[result_df['Asset - Tag'].isnull(), ['Remark']]= 'No Match'
                result_df.loc[result_df['Task Status']=='Closed', ['Remark']]= 'Closed'
                result_df.drop('_merge', axis=1, inplace=True)
                sub_df=result_df[['Report/RFI No', 'Parameter Reading 01', 'Parameter Reading 02', 'Parameter Reading 03', 'Parameter Reading 04', 'Parameter Reading 05', 'Parameter Reading 06', 'Parameter Reading 07', 'Parameter Reading 08', 'Parameter Reading 09', 'Parameter Reading 10', 'Completed Date', 'Completed By', 'Remark']].copy()

                 # Create a BytesIO object for each Excel file
                excel_file1 = io.BytesIO()
                excel_file2 = io.BytesIO()
                excel_file3 = io.BytesIO()
                excel_file4 = io.BytesIO()
                excel_file5 = io.BytesIO()

               # Write the DataFrames to the BytesIO objects
                with pd.ExcelWriter(excel_file1, engine='openpyxl') as writer:
                    sub_df.to_excel(writer, index=False, sheet_name='Sheet1')
                    apply_formatting(writer, 'Sheet1', result_df)
                
                result_df=result_df.loc[(result_df['Remark']=='Good Match') & (result_df['Task Status']!='Closed')]
                PNT_005_A_Import_SCDB_1 =result_df.copy()
                PNT_005_A_Import_SCDB_1.columns=['Task ID', 'Task - Name', 'Asset - Tag', 'Task Status', 'Step Sequence', 'Step Point', 'Inspection Type', 'Report/RFI No', 'Parameter Reading 01', 'Parameter Reading 02', 'Parameter Reading 03', 'Parameter Reading 04', 'Parameter Reading 05', 'Parameter Reading 06', 'Parameter Reading 07', 'Parameter Reading 08', 'Parameter Reading 09', 'Parameter Reading 10', 'Completed Date', 'Completed By', 'Remark']
                PNT_005_A_Import_SCDB_1['Step Sequence']=1
                PNT_005_A_Import_SCDB_1['Step Point']=1
                PNT_005_A_Import_SCDB_1.drop(['Task - Name','Remark','Asset - Tag','Task Status','Report/RFI No'], axis=1, inplace=True)
                PNT_005_A_Import_SCDB_1.columns=['Task - Name', 'Step Sequence', 'Step Point', 'Inspection Type',
                                                    'Parameter Reading 01', 'Parameter Reading 02', 'Parameter Reading 03',
                                                    'Parameter Reading 04', 'Parameter Reading 05', 'Parameter Reading 06',
                                                    'Parameter Reading 07', 'Parameter Reading 08', 'Parameter Reading 09',
                                                    'Parameter Reading 10', 'Completed Date', 'Completed By']
                PNT_005_A_Import_SCDB_1.to_excel(excel_file2, index=False, header=True)
                Link_doc_ITR_col=result_df[['Task ID','Task Model (Name)','Report/RFI No']].copy()
                Link_doc_ITR_col.columns=['Task ID','Task Model (Name)','Documents (Name List)']
                Link_doc_ITR_col.to_excel(excel_file3, index=False, header=True)
                Create_Document=result_df[['Report/RFI No']].copy()
                Create_Document['Document - Revision']=0
                Create_Document.columns=['Document - Name/ID','Document - Revision']
                Create_Document.drop_duplicates(inplace=True)
                Create_Document.to_excel(excel_file4, index=False, header=True)
                pnt05_Status_1=result_df[['Task ID','Task Model (Name)','Parameter Reading 08','Completed By']].copy()          
                today_date = datetime.today().strftime('%m/%d/%Y')
                # Convert to string (although it's already a string due to strftime)
                today_date_str = str(today_date)
                pnt05_Status_1[['Submitted Date']]=today_date_str
                pnt05_Status_1[['Actual End Date']]=today_date_str
                pnt05_Status_1[['Actual Start Date']]=today_date_str
                pnt05_Status_1=pnt05_Status_1[['Task ID', 'Task Model (Name)', 'Submitted Date','Parameter Reading 08', 'Actual End Date','Completed By','Actual Start Date']]
                pnt05_Status_1.columns=['Task ID', 'Task Model (Name)', 'Submitted Date','Submitted By', 'Actual End Date','Completed By','Actual Start Date']
                pnt05_Status_1.to_excel(excel_file5, index=False, header=True)
                # Create a BytesIO object for the zip file
                zip_file = io.BytesIO()

                # Write the BytesIO objects to the zip file
                with zipfile.ZipFile(zip_file, 'w') as zipf:
                    zipf.writestr("Subcon.xlsx", excel_file1.getvalue())
                    zipf.writestr("PNT_005_A_Import_SCDB.xlsx", excel_file2.getvalue())
                    zipf.writestr("Link_doc_ITR.xlsx", excel_file3.getvalue())
                    zipf.writestr("Create_Document.xlsx", excel_file4.getvalue())
                    zipf.writestr("pnt05_Status.xlsx", excel_file5.getvalue())

                # Provide a download link for the zip file
                st.download_button(
                    label="Download All Files",
                    data=zip_file.getvalue(),
                    file_name='all_files_003004.zip',
                    mime='application/zip'
                )
                
            except Exception as e:
                st.error(f"Error executing logic and generating Excel files: {e}")
        
        else:
            st.warning("Please upload all files before executing logic.")
