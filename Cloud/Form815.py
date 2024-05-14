import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import gc 
import io
import zipfile
import warnings
warnings.filterwarnings('ignore')

def app():
    # Streamlit app setup
    st.markdown("<h1 style='text-align: center; width: 100%;'>Form-815 File Uploader/Downloader</h1>", unsafe_allow_html=True)
    # Add instruction for refreshing the page
    st.markdown("<p style='text-align: left; color: blue; font-family: Arial;'>For refresh and remove history, refresh from Browser right top most corner.</p>", unsafe_allow_html=True)

    # Function to browse and select a file
    def browse_file(file_type):
        file_path = st.file_uploader(f"Upload {file_type} File:", type=["xlsx", "xls"], key=f"{file_type.lower().replace(' ', '_')}")
        if file_path:
            st.success(f"{file_type} File uploaded successfully!")
            return file_path

    # Browse and upload files
    file_scdb_path = browse_file("SCDB")
    file_subcontractor_path = browse_file("Subcontractor")
    file_scdb_mapping_path = browse_file("SCDB Name Mapping")
    file_template_path = browse_file("Template")

    # This function formatting for the subcontractor report
    def apply_formatting(writer, sheet_name, dataframe):
            # Access the worksheet
            ws = writer.sheets[sheet_name]
    
            # Define the fill pattern for the red color
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
            # Iterate through each row in the DataFrame and check the condition
            for row_number, row in dataframe.iterrows():
                # Check if the condition is met in the 'Remark' column
                if any(keyword in row['Remark'] for keyword in ['Task_ID and Tag No. not matched', 'Task_ID not matched', 'Tag No. NO not matched', 'Closed']):
                    # Apply the red fill to the entire row
                    for col_number, value in enumerate(row):
                        ws.cell(row=row_number + 2, column=col_number + 1).fill = red_fill

    # Missing values in Various Name columns after Mapping with Dictonary 
    def apply_formatting_missing(writer, sheet_name, dataframe, columns_to_check, fill_color):
        # Access the worksheet
        ws = writer.sheets[sheet_name]

        # Define the fill pattern
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Iterate through each row in the DataFrame
        for row_number, row in dataframe.iterrows():
            # Check if any of the specified columns have missing values (NaN)
            if any(pd.isnull(row[col]) for col in columns_to_check):
                # Apply the fill color to the entire row
                for col_number, value in enumerate(row):
                    ws.cell(row=row_number + 2, column=col_number + 1).fill = fill
    
    # Define for Accept/Reject and N/A
    def apply_conditional_formatting(sheet, column_name, valid_values, fill_color="FF0000"):
        # Get the column index from the column name
        column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == column_name:
                    column_index = cell.column
                    break
            if column_index:
                break

        # If column index is found, iterate through cells in the specified column to apply conditional formatting
        if column_index:
            for row in sheet.iter_rows(min_row=2, min_col=column_index, max_row=sheet.max_row, max_col=column_index):
                for cell in row:
                    if cell.value not in valid_values:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        else:
            print(f"Column '{column_name}' not found in the worksheet.")


    # Execute logic and download
    if st.button("Execute Program"):

        if file_scdb_path and file_subcontractor_path and file_scdb_mapping_path and file_template_path:
            try:
                st.info("Executing Programme for files generation ...")
            
                # Load dataframes
                # Master file downloaded from SCDB
                selected_columns = ['Task ID', 'Task Model (Name)', 'Asset - Tag', 'Task State']
                df_master = pd.read_excel(file_scdb_path, dtype=str, usecols=selected_columns)
                df_master.columns = ['Task ID', 'Task Model (Name)', 'Tag No.', 'Task State']
                
                # Subcon Data
                df_subcon=pd.read_excel(file_subcontractor_path,keep_default_na=False,skiprows=1,dtype=str,sheet_name='Sheet1')
                result_df = pd.merge(df_master, df_subcon, how='right', indicator=True, on=['Task ID', 'Tag No.'])
                not_matching_rows = result_df.loc[result_df['_merge'] == 'right_only', ['Task ID', 'Tag No.']]
                result_df['Remark'] = 'Good Match'
                result_df.loc[result_df['Task State'] == 'Closed', 'Remark'] = 'Closed'
                
                for index, row in not_matching_rows.iterrows():
                    id_match = df_master['Task ID'].isin([row['Task ID']])
                    name_match = df_master['Tag No.'].isin([row['Tag No.']])
                    if (not id_match.any()) & (not name_match.any()):
                        result_df.loc[(result_df['Task ID'] == row['Task ID']) & (result_df['Tag No.'] == row['Tag No.']), 'Remark'] = 'Task_ID and Tag No. not matched'
                    elif not id_match.any():
                        result_df.loc[(result_df['Task ID'] == row['Task ID']) & (result_df['Tag No.'] == row['Tag No.']), 'Remark'] = 'Task_ID not matched'
                    else:
                        result_df.loc[(result_df['Task ID'] == row['Task ID']) & (result_df['Tag No.'] == row['Tag No.']), 'Remark'] = 'Tag No. NO not matched'

                result_df.drop('_merge',axis=1,inplace=True)
                result_df[['Initials / Date','Submitted Date','Completed Date']] = result_df[['Initials / Date','Submitted Date','Completed Date']].apply(pd.to_datetime, errors='coerce')
                result_df[['Initials / Date','Submitted Date','Completed Date']] = result_df[['Initials / Date','Submitted Date','Completed Date']].apply(lambda x: x.dt.strftime('%d-%b-%y'))
                
                date_columns = ['Initials / Date','Submitted Date','Completed Date']
                columns_to_drop = ['Task Model (Name)', 'Task State']
                result=df_subcon.copy()
                result['Remark'] = result_df['Remark']
                
                # Convert to datetime and format as 'dd-mmm-yy'
                result[date_columns] = result[date_columns].apply(pd.to_datetime, errors='coerce').apply(lambda x: x.dt.strftime('%d-%b-%y'))

                # Convert the Excel file to bytes
                excel_file = io.BytesIO()
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    result.to_excel(writer, index=False, sheet_name='Sheet1')

                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']

                    # Apply the custom date format for specific columns
                    date_format = 'DD-MMM-YY'
                    for col in date_columns:
                        col_index = result.columns.get_loc(col)
                        for row in range(2, worksheet.max_row + 1):  # Assuming header is in the first row
                            worksheet.cell(row=row, column=col_index + 1).number_format = date_format

                    apply_formatting(writer, 'Sheet1', result)

                excel_file.seek(0)

                # List of values to exclude
                exclude_values = ['Closed','Task_ID and Tag No. not matched', 'Task_ID not matched', 'Tag No. NO not matched']
                # Filtering rows based on the 'Remark' column
                result_df = result_df[~result_df['Remark'].isin(exclude_values)]
                result_df=result_df.reset_index(drop=True)

                #Creating the empty Data Frame with Header for Step-1
                col_names_as_output = pd.read_excel(file_template_path, nrows=0).columns.tolist()
                Step_1=pd.DataFrame(columns=col_names_as_output)

                # Creating Data Frame as per Step-2
                columns = ['Task ID', 'Task Model (Name)', 'Actual Start Date', 'Submitted Date', 'Submitted By', 'Actual End Date', 'Completed By']

                # Create a blank DataFrame
                Step2_df = pd.DataFrame(columns=columns)
                Temp=pd.read_excel(file_template_path)

                for index, _ in result_df.iterrows():
                    Temp_update=Temp.copy()
                    Temp_update['Completed Date'] = result_df.loc[index, 'Completed Date']
                    Temp_update['Completed By'] = result_df.loc[index, 'Submitted By SCTR']
                    Temp_update['Task - Name'] = result_df.loc[index,'Task ID']
                    Temp_update.at[0,'Inspection Answer']=result_df.at[index,'Drawing No:']
                    Temp_update.at[1,'Step Answer']=result_df.at[index,'Cable tag/type/size/rating are as per the cable schedule.\n(Acc / Rej / NA)']
                    Temp_update.at[2,'Step Answer']=result_df.at[index,'Cables neatly and securely arranged, supported without stress on the termination, no sharp bends, and no damages to outer sheath/conductors\n(Acc / Rej / NA)']
                    Temp_update.at[3,'Step Answer']=result_df.at[index,'Cable of different voltage levels and / or service shall be segregated.\n(Acc / Rej / NA)']
                    Temp_update.at[4,'Step Answer']=result_df.at[index,'Check trefoil formation of single core cables supplying three phase loads.\n(Acc / Rej / NA)']
                    Temp_update.at[5,'Step Answer']=result_df.at[index,'Verify that cable gland type and size are correct/suitable and as per cable schedule.\n(Acc / Rej / NA)']
                    Temp_update.at[6,'Step Answer']=result_df.at[index,'All gland components are correctly assembled and tightened.\n(Acc / Rej / NA)']
                    Temp_update.at[7,'Step Answer']=result_df.at[index,'Verify that correct IP/ sealing washers are fitted to cable gland (where required)\n(Acc / Rej / NA)']
                    Temp_update.at[8,'Step Answer']=result_df.at[index,'Check termination lug are correct size/type as per specifications\n(Acc / Rej / NA)']
                    Temp_update.at[9,'Step Answer']=result_df.at[index,'Check termination at both ends is as per phase, color code and identification\n(Acc / Rej / NA)']
                    Temp_update.at[10,'Step Answer']=result_df.at[index,'Earth bonding of cables is satisfactory and as per drawings.\n(Acc / Rej / NA)']
                    Temp_update.at[11,'Step Answer']=result_df.at[index,'All electrical connections are tightened, insulated and tape (torquing if required).\n(Acc / Rej / NA)']
                    Temp_update.at[12,'Step Answer']=result_df.at[index,'All Spare cables are correctly terminated.\n(Acc / Rej / NA)']

                    Step_1 = pd.concat([Step_1, Temp_update], ignore_index=True)
                    Step2_L=[None]*7
                    Step2_L[0]=result_df.loc[index,'Task ID']
                    Step2_L[1]=result_df.loc[index,'Task Model (Name)']
                    Step2_L[2]=result_df.loc[index,'Initials / Date']
                    Step2_L[3]=result_df.loc[index,'Submitted Date']
                    Step2_L[4]=result_df.loc[index,'Submitted By SCTR']
                    Step2_L[5]=result_df.loc[index,'Completed Date']
                    Step2_L[6]=result_df.loc[index,'Completed By CTR']
                    # Create the new row to a DataFrame
                    new_row_df = pd.DataFrame([Step2_L], columns=columns)
                    Step2_df = pd.concat([Step2_df, new_row_df], ignore_index=True)
                Step_1 = Step_1.astype(str)
                Step_1 = Step_1.replace('nan', '')
                # Save DataFrame to BytesIO object
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    Step_1.to_excel(writer, sheet_name='Sheet1', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    # Apply the custom date format for specific columns
                    date_format = 'DD-MMM-YY'
                    date_columns = ['Completed Date']
                    for col in date_columns:
                        col_index = Step_1.columns.get_loc(col)
                        for row in range(2, worksheet.max_row + 1):  # Assuming header is in the first row
                            worksheet.cell(row=row, column=col_index + 1).number_format = date_format
                output.seek(0)  # go to the start of the stream

                # Load the workbook
                workbook = load_workbook(output)
                sheet = workbook.active

                # Define the column names
                column_names = ['Completed By']
 
                # Get the column indices from the column names
                column_indices = {}
                for column_name in column_names:
                    for col in sheet.iter_cols(min_row=1, max_row=1):
                        for cell in col:
                            if cell.value == column_name:
                                column_indices[column_name] = cell.column
                                break
 
                # Skip if any column not found
                for column_name in column_names:
                    if column_name not in column_indices:
                        print(f"Column '{column_name}' not found in the worksheet.")
                        continue
 
                # Get the column index for 'Completed By' column
                completed_by_column_index = column_indices['Completed By']
 
                # Iterate over each row
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
                    completed_by_cell = sheet.cell(row=row_num, column=completed_by_column_index)  # Get the 'Completed By' cell
 
                    # Check if 'Completed By' cell is empty
                    if completed_by_cell.value is None:
                        completed_by_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                apply_conditional_formatting(sheet, 'Step Answer', ['Accept', 'Reject', 'N/A', None])

                    # Save the workbook to a new BytesIO object
                output = io.BytesIO()
                workbook.save(output)
                output.seek(0)
                workbook = None
                gc.collect()  # call the garbage collector
                    # New DataFrame for SCDB Name Mapping
                mapping_file = pd.read_excel(file_scdb_mapping_path)
                # Create a preprocessed dictionary with case-insensitive keys and removed white spaces
                preprocessed_dictionary = dict(
                    zip(
                        map(lambda x: "".join(str(x).lower().split()), mapping_file.iloc[:, 0]),
                        mapping_file.iloc[:, 1]
                    )
                )
                # Replace the 'Completed By' column with the preprocessed values
                def preprocess_column(column):
                    return (
                        column
                        .map(lambda x: "".join(str(x).lower().split()))
                        .map(preprocessed_dictionary)
                    )

                Step2_df['Completed By'] = preprocess_column(Step2_df['Completed By'])
                Step2_df['Submitted By'] = preprocess_column(Step2_df['Submitted By'])

                # Convert the Excel file to bytes
                excel_file2 = io.BytesIO()
                with pd.ExcelWriter(excel_file2, engine='openpyxl') as writer:
                    Step2_df.to_excel(writer, index=False, sheet_name='Sheet1')
                    # Call the function with specific columns and red fill color
                    apply_formatting_missing(writer, 'Sheet1', Step2_df, ['Submitted By', 'Completed By'], 'FF0000')


                excel_file2.seek(0)
                st.success("Files generated successfully!")
                st.write("Download your files here:")
                zip_file = io.BytesIO()
                with zipfile.ZipFile(zip_file, 'w') as zipf:
                    zipf.writestr("Report.xlsx", excel_file.getvalue())
                    zipf.writestr("Step-1.xlsx", output.getvalue())
                    zipf.writestr("Step-2.xlsx", excel_file2.getvalue())

                # Provide a download link for the zip file
                st.download_button(
                    label="Download All Files",
                    data=zip_file,
                    file_name='all_files_813.zip',
                    mime='application/zip'
                )
                
            except Exception as e:
                st.error(f"Error executing logic and generating Excel files: {e}")
        
        else:
            st.warning("Please upload all files before executing logic.")
