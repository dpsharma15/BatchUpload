import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import gc 
import io
import zipfile
import warnings
warnings.filterwarnings('ignore')
from collections import OrderedDict
import pandas as pd
from difflib import get_close_matches

def app():
    # Streamlit app setup
    st.markdown("<h1 style='text-align: center; width: 100%;'>Form-308 File Uploader/Downloader</h1>", unsafe_allow_html=True)
    # Add instruction for refreshing the page
    st.markdown("<p style='text-align: left; color: blue; font-family: Arial;'>For refresh and remove history, refresh from Browser right top most corner.</p>", unsafe_allow_html=True)

    # Function to browse and select a file
    def browse_file(file_type):
        file_path = st.file_uploader(f"Upload {file_type} File:", type=["xlsx", "xls"], key=f"{file_type.lower().replace(' ', '_')}")
        if file_path:
            st.success(f"{file_type} File uploaded successfully!")
            return file_path

    # Browse and upload files
    file_scdb_path = browse_file("SCDB")
    file_subcontractor_path = browse_file("Subcontractor")
    file_scdb_mapping_path = browse_file("SCDB Name Mapping")

    # This function formatting for the subcontractor report
    def apply_formatting(writer, sheet_name, dataframe):
            # Access the worksheet
            ws = writer.sheets[sheet_name]
    
            # Define the fill pattern for the red color
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
            # Iterate through each row in the DataFrame and check the condition
            for row_number, row in dataframe.iterrows():
                # Check if the condition is met in the 'Remark' column
                if any(keyword in row['Remark'] for keyword in ['Task_ID and TAG/ID NO not matched', 'Task_ID not matched', 'TAG/ID NO not matched', 'Closed']):
                    # Apply the red fill to the entire row
                    for col_number, value in enumerate(row):
                        ws.cell(row=row_number + 2, column=col_number + 1).fill = red_fill

    # Missing values in Various Name columns after Mapping with Dictonary 
    def apply_formatting_missing(writer, sheet_name, dataframe, columns_to_check, fill_color):
        # Access the worksheet
        ws = writer.sheets[sheet_name]

        # Define the fill pattern
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Iterate through each row in the DataFrame
        for row_number, row in dataframe.iterrows():
            # Check if any of the specified columns have missing values (NaN)
            if any(pd.isnull(row[col]) for col in columns_to_check):
                # Apply the fill color to the entire row
                for col_number, value in enumerate(row):
                    ws.cell(row=row_number + 2, column=col_number + 1).fill = fill

    # Finally downlaod the different files
    def create_download_button(excel_file, label, file_name):
        return st.download_button(
            label=label,
            data=excel_file,
            file_name=file_name,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    # Define for Accept/Reject and N/A
    def apply_conditional_formatting(sheet, column_name, valid_values, fill_color="FF0000"):
        # Get the column index from the column name
        column_index = None
        for col in sheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == column_name:
                    column_index = cell.column
                    break
            if column_index:
                break

        # If column index is found, iterate through cells in the specified column to apply conditional formatting
        if column_index:
            for row in sheet.iter_rows(min_row=2, min_col=column_index, max_row=sheet.max_row, max_col=column_index):
                for cell in row:
                    if cell.value not in valid_values:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        else:
            print(f"Column '{column_name}' not found in the worksheet.")

    # Execute logic and download
    if st.button("Execute Program"):

        if file_scdb_path and file_subcontractor_path and file_scdb_mapping_path:
            try:
                st.info("Executing Programme for files generation ...")
        
                # Master file downloaded from SCDB
                selected_columns = ['Task ID', 'Task Model (Name)', 'Asset - Tag', 'Task Status']
                df_master = pd.read_excel(file_scdb_path, dtype=str, usecols=selected_columns)
                df_master.columns = ['Taskid ', 'Task Model (Name)', 'TAG/ID NO', 'Task Status']

                # Subcon Data
                df_subcon = pd.read_excel(file_subcontractor_path, dtype=str, skiprows=1)
                for col in ['Batch time', 'Pouring Start Time', 'Pouring Finish Time']:
                    df_subcon[col] = df_subcon[col].map(lambda s: str(s)[0:5])

                df_subcon['Submitted Date\n(mm/dd/yyyy)'] = df_subcon['Submitted Date\n(mm/dd/yyyy)'].apply(
                    lambda x: pd.to_datetime(x, errors='coerce', format='%Y-%m-%d %H:%M:%S') if isinstance(x, str) and len(x) == 19 else pd.to_datetime(x, errors='coerce'))

                df_subcon['Submitted Date\n(mm/dd/yyyy)'] = df_subcon['Submitted Date\n(mm/dd/yyyy)'].dt.strftime('%d-%b-%y')

                result_df = pd.merge(df_master, df_subcon, how='right', indicator=True, on=['Taskid ', 'TAG/ID NO'])

                not_matching_rows = result_df.loc[result_df['_merge'] == 'right_only', ['Taskid ', 'TAG/ID NO']]
                
                result_df['Remark'] = 'Good Match'

                for index, row in not_matching_rows.iterrows():
                    id_match = df_master['Taskid '].isin([row['Taskid ']])
                    name_match = df_master['TAG/ID NO'].isin([row['TAG/ID NO']])

                    if (not id_match.any()) & (not name_match.any()):
                        result_df.loc[(result_df['Taskid '] == row['Taskid ']) & (result_df['TAG/ID NO'] == row['TAG/ID NO']), 'Remark'] = 'Task_ID and TAG/ID NO not matched'

                    elif not id_match.any():
                        result_df.loc[(result_df['Taskid '] == row['Taskid ']) & (result_df['TAG/ID NO'] == row['TAG/ID NO']), 'Remark'] = 'Task_ID not matched'

                    else:
                        result_df.loc[(result_df['Taskid '] == row['Taskid ']) & (result_df['TAG/ID NO'] == row['TAG/ID NO']), 'Remark'] = 'TAG/ID NO not matched'

                result_df.loc[result_df['Task Status'] == 'Closed', 'Remark'] = result_df['Remark'] + result_df['Task Status']
                result_df_both=result_df[result_df['_merge']=='both'] ## For step1
                result_df.drop('_merge', axis=1, inplace=True)

                # Convert the Excel file to bytes
                excel_file = io.BytesIO()
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                    apply_formatting(writer, 'Sheet1', result_df)
                excel_file.seek(0)

                # List of values to exclude
                exclude_values = ['Task_ID and TAG/ID NO not matched', 'Task_ID not matched', 'TAG/ID NO not matched', 'Closed']
                # Filtering rows based on the 'Remark' column
                result_df = result_df[~result_df['Remark'].isin(exclude_values)]
                result_df =result_df_both.drop('_merge', axis=1)
                result_df=result_df.reset_index(drop=True)

                # Your column mapping dictionary
                column_mapping = {
                        'Taskid ': 'Task - Name',
                        'Drawing/Doc No': 'Parameter Reading 01',
                        'Truck No': 'Parameter Reading 02',
                        'Ticket No': 'Parameter Reading 03',
                        'Concrete Volume Supplied': 'Parameter Reading 04',
                        'Cumulative Volume supplied': 'Parameter Reading 05',
                        'Batch time': 'Parameter Reading 06',
                        'Pouring Start Time': 'Parameter Reading 07',
                        'Pouring Finish Time': 'Parameter Reading 08',
                        'Slump (mm)': 'Parameter Reading 09',
                        'Temprature (⁰C)': 'Parameter Reading 10',
                        'Number of cubes made': 'Parameter Reading 11',
                        'Submitted Date\n(mm/dd/yyyy)': 'Completed Date',
                        'Submitted By': 'Completed By',
                        'Rev no': 'Rev no',
                        'Task Model (Name)':'Task Model (Name)',  
                        'Completed By\n':'Completed By_step2'
                }
    
                # Filter columns that are present in the column_mapping dictionary
                columns_to_rename = [col for col in result_df.columns if col in column_mapping]

                # Rename selected columns using the column_mapping dictionary
                result_df.rename(columns={col: column_mapping.get(col, col) for col in columns_to_rename}, inplace=True)
                new_column_names = list(column_mapping.values())
                result_df=result_df[new_column_names]

                # required to generate the file as o/p, Extra column getting updated
                result_df['Step Sequence'] = ''
                result_df['Step Point'] = ''
                result_df['Inspection Type'] = ''
                result_df['Step Action'] = ''

                # Assuming result_df is our DataFrame as per required step-1 excel file ,one additional column for rev
                result_df = result_df[['Task - Name', 'Step Sequence', 'Step Point', 'Inspection Type',
                                    'Step Action', 'Parameter Reading 01', 'Parameter Reading 02',
                                    'Parameter Reading 03', 'Parameter Reading 04', 'Parameter Reading 05',
                                    'Parameter Reading 06', 'Parameter Reading 07', 'Parameter Reading 08',
                                    'Parameter Reading 09', 'Parameter Reading 10', 'Parameter Reading 11',
                                    'Completed Date', 'Completed By','Rev no','Task Model (Name)','Completed By_step2']]
                
                #column List required for step-1
                col=['Task - Name', 'Step Sequence', 'Step Point', 'Inspection Type',
                    'Step Action', 'Parameter Reading 01', 'Parameter Reading 02',
                    'Parameter Reading 03', 'Parameter Reading 04', 'Parameter Reading 05',
                    'Parameter Reading 06', 'Parameter Reading 07', 'Parameter Reading 08',
                    'Parameter Reading 09', 'Parameter Reading 10', 'Parameter Reading 11',
                    'Completed Date', 'Completed By']
                df_final_to_write = pd.DataFrame(columns=col)

                # Initialize an OrderedDict to keep track of order and count
                order_and_count = OrderedDict()

                L = [None] * 18

                # Iterate over the values in the series while maintaining order
                for index,value in enumerate(result_df['Task - Name']):
                    # Update the count for the value
                    order_and_count[value] = order_and_count.get(value, 0) + 1
                    
                    # Check the count for each value
                    count = order_and_count[value]
                    # Perform different tasks based on repetition
                    if count > 1:
                        # Task for repeated value
                        L[0]=result_df.loc[index,'Task - Name']
                        L[1]=count+1
                        L[2]=count
                        L[5]=count
                        for j in range(6,18):
                            L[j]=result_df.iloc[index,j]
                        new_row = pd.Series(L, index=df_final_to_write.columns)
                        df_final_to_write = pd.concat([df_final_to_write, new_row.to_frame().T], ignore_index=True)
                        L = [None] * 18
                        
                #         print(f"Task for repeated value: {value}")
                    else:
                        # Task for non-repeated value
                        L[0]=result_df.loc[index,'Task - Name']
                        L[1]=count
                        L[2]=0.1
                        L[5]=result_df.loc[index,'Parameter Reading 01']
                        L[6]=result_df.loc[index,'Rev no']
                        L[16]=result_df.loc[index,'Completed Date']
                        L[17]=result_df.loc[index,'Completed By']
                        new_row = pd.Series(L, index=df_final_to_write.columns)
                        df_final_to_write = pd.concat([df_final_to_write, new_row.to_frame().T], ignore_index=True)
                        L = [None] * 18
                        L[0]=result_df.loc[index,'Task - Name']
                        L[1]=count+1
                        L[2]=count
                        L[5]=count
                        for j in range(6,18):
                            L[j]=result_df.iloc[index,j]
                        new_row = pd.Series(L, index=df_final_to_write.columns)
                        df_final_to_write = pd.concat([df_final_to_write, new_row.to_frame().T], ignore_index=True)
                        L = [None] * 18
                        # Create a BytesIO buffer to hold the Excel file content
                excel_buffer = io.BytesIO()

                # Write the DataFrame to an Excel file in the BytesIO buffer
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df_final_to_write.to_excel(writer, index=False, sheet_name='Sheet1')

                col_step_2=['Task - Name','Task Model (Name)','Completed Date','Completed Date','Completed By','Completed Date','Completed By_step2']
                df_step2=result_df[col_step_2]
                df_step2.columns=['Task ID', 'Task Model (Name)', 'Actual Start Date', 'Submitted Date',
                'Submitted By', 'Actual End Date', 'Completed By']
                # Drop duplicates based on 'Task ID'
                df_step2 = df_step2.drop_duplicates(subset='Task ID', keep='first')
                # New DataFrame for SCDB Name Mapping
                def find_closest_match(value, dictionary):
                    matches = get_close_matches(value, dictionary.keys(), n=1, cutoff=0.6)
                    return dictionary.get(matches[0]) if matches else None

                # Read the mapping file into a DataFrame
                mapping_file = pd.read_excel(file_scdb_mapping_path)

                # Create a preprocessed dictionary with case-insensitive keys and removed white spaces
                preprocessed_dictionary = dict(
                    zip(
                        map(lambda x: "".join(str(x).lower().split()), mapping_file.iloc[:, 0]),
                        mapping_file.iloc[:, 1]
                    )
                )

                # Your DataFrame series (df_step2['Completed By'] in this example)
                # Replace the 'Completed By' column with the preprocessed values using partial match
                df_step2['Completed By'] = df_step2['Completed By'].apply(lambda x: find_closest_match(''.join(str(x).lower().split()), preprocessed_dictionary))

                # Convert the Excel file to bytes
                excel_file2 = io.BytesIO()
                with pd.ExcelWriter(excel_file2, engine='openpyxl') as writer:
                    df_step2.to_excel(writer, index=False, sheet_name='Sheet1')
                    # Call the function with specific columns and red fill color
                    apply_formatting_missing(writer, 'Sheet1', df_step2, ['Completed By'], 'FF0000')

                excel_file2.seek(0)
                st.success("Files generated successfully!")
                st.write("Download your files here:")
                zip_file = io.BytesIO()
                with zipfile.ZipFile(zip_file, 'w') as zipf:
                    zipf.writestr("Report.xlsx", excel_file.getvalue())
                    zipf.writestr("Step-1.xlsx", excel_buffer.getvalue())
                    zipf.writestr("Step-2.xlsx", excel_file2.getvalue())

                # Provide a download link for the zip file
                st.download_button(
                    label="Download All Files",
                    data=zip_file,
                    file_name='all_files_308.zip',
                    mime='application/zip'
                )
                
            except Exception as e:
                st.error(f"Error executing logic and generating Excel files: {e}")
        
        else:
            st.warning("Please upload all files before executing logic.")