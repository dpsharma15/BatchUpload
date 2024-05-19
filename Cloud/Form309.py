import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill
import io
import zipfile
import warnings
warnings.filterwarnings('ignore')
import pandas as pd


def app():
    # Streamlit app setup
    st.markdown("<h1 style='text-align: center; width: 100%;'>Form-309 File Uploader/Downloader</h1>", unsafe_allow_html=True)
    # Add instruction for refreshing the page
    st.markdown("<p style='text-align: left; color: blue; font-family: Arial;'>For refresh and remove history, refresh from Browser Left top most corner.</p>", unsafe_allow_html=True)

    # Function to browse and select a file
    def browse_file(file_type):
        file_path = st.file_uploader(f"Upload {file_type} File:", type=["xlsx", "xls"], key=f"{file_type.lower().replace(' ', '_')}")
        if file_path:
            st.success(f"{file_type} File uploaded successfully!")
            return file_path

    # Browse and upload files
    file_scdb_path = browse_file("SCDB")
    file_subcontractor_path = browse_file("Subcontractor")
    file_scdb_mapping_path = browse_file("SCDB Name Mapping")
    file_template_path = browse_file("Upload Template")

    # This function formatting for the subcontractor report
    def apply_formatting(writer, sheet_name, dataframe):
            # Access the worksheet
            ws = writer.sheets[sheet_name]
    
            # Define the fill pattern for the red color
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
            # Iterate through each row in the DataFrame and check the condition
            for row_number, row in dataframe.iterrows():
                # Check if the condition is met in the 'Remark' column
                if any(keyword in row['Remark'] for keyword in ['Closed','Task_ID and Tag No. not matched', 'Task_ID not matched', 'Tag No. NO not matched']):
                    # Apply the red fill to the entire row
                    for col_number, value in enumerate(row):
                        ws.cell(row=row_number + 2, column=col_number + 1).fill = red_fill

    # Missing values in Various Name columns after Mapping with Dictonary 
    def apply_formatting_missing(writer, sheet_name, dataframe, columns_to_check, fill_color):
        # Access the worksheet
        ws = writer.sheets[sheet_name]

        # Define the fill pattern
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        # Iterate through each row in the DataFrame
        for row_number, row in dataframe.iterrows():
            # Check if any of the specified columns have missing values (NaN)
            if any(pd.isnull(row[col]) for col in columns_to_check):
                # Apply the fill color to the entire row
                for col_number, value in enumerate(row):
                    ws.cell(row=row_number + 2, column=col_number + 1).fill = fill

 
    if st.button("Execute Program"):

        if file_scdb_path and file_subcontractor_path and file_scdb_mapping_path:
            try:
                st.info("Executing Programme for files generation ...")
        
                # Master file downloaded from SCDB
                selected_columns = ['Task ID', 'Task Model (Name)', 'Asset - Tag', 'Task State']
                df_master = pd.read_excel(file_scdb_path, dtype=str, usecols=selected_columns)
                df_master.columns = ['Task ID', 'Task Model (Name)', 'Tag No.', 'Task State']
            
                df_subcon = pd.read_excel(file_subcontractor_path, keep_default_na=False, skiprows=1, dtype=str)
                result_df = pd.merge(df_master, df_subcon, how='right', indicator=True, on=['Task ID', 'Tag No.'])
                not_matching_rows = result_df.loc[result_df['_merge'] == 'right_only', ['Task ID', 'Tag No.']]
                result_df['Remark'] = 'Good Match'
                result_df.loc[result_df['Task State'] == 'Closed', 'Remark'] = 'Closed'
            
                for index, row in not_matching_rows.iterrows():
                    id_match = df_master['Task ID'].isin([row['Task ID']])
                    name_match = df_master['Tag No.'].isin([row['Tag No.']])
                    if (not id_match.any()) & (not name_match.any()):
                        result_df.loc[(result_df['Task ID'] == row['Task ID']) & (result_df['Tag No.'] == row['Tag No.']), 'Remark'] = 'Task_ID and Tag No. not matched'
                    elif not id_match.any():
                        result_df.loc[(result_df['Task ID'] == row['Task ID']) & (result_df['Tag No.'] == row['Tag No.']), 'Remark'] = 'Task_ID not matched'
                    else:
                        result_df.loc[(result_df['Task ID'] == row['Task ID']) & (result_df['Tag No.'] == row['Tag No.']), 'Remark'] = 'Tag No. NO not matched'
                result_df.drop('_merge', axis=1, inplace=True)
                result_df[['Initials / Date','Submitted Date','Completed Date','Cal Expiry']] = result_df[['Initials / Date','Submitted Date','Completed Date','Cal Expiry']].apply(pd.to_datetime, errors='coerce')
                result_df[['Initials / Date','Submitted Date','Completed Date']] = result_df[['Initials / Date','Submitted Date','Completed Date']].apply(lambda x: x.dt.strftime('%d-%b-%y'))
                result_df[['Cal Expiry']]=result_df[['Cal Expiry']].apply(lambda x: x.dt.strftime('%d-%b-%Y'))
            
                date_columns = ['Initials / Date','Submitted Date','Completed Date','Closed Date','Cal Expiry']
                columns_to_drop = ['Task ID', 'Task Model (Name)', 'Tag No.', 'Task State']
                result=result_df.drop(columns=columns_to_drop)
            
                excel_file1 = io.BytesIO()
                with pd.ExcelWriter(excel_file1, engine='openpyxl') as writer:
                    result.to_excel(writer, index=False, sheet_name='Sheet1')
                    # workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    date_format = 'DD-MMM-YY'
                    for col in date_columns:
                        for col_number, header in enumerate(result.columns, start=1):
                            if header == col:
                                for row in worksheet.iter_rows(min_row=2, max_col=result.shape[1], max_row=worksheet.max_row):
                                    for cell in row:
                                        if cell.column == col_number:
                                            cell.number_format = date_format
                    apply_formatting(writer, 'Sheet1', result)
                excel_file1.seek(0)

                # Further processing...
                exclude_values = ['Closed','Task_ID and Tag No. not matched', 'Task_ID not matched', 'Tag No. NO not matched']
                result_df = result_df[~result_df['Remark'].isin(exclude_values)]
                result_df = result_df.reset_index(drop=True)
            
                col_names_as_output = pd.read_excel(file_template_path, nrows=0).columns.tolist()
                Step_1 = pd.DataFrame(columns=col_names_as_output)
            
                columns = ['Task ID', 'Task Model (Name)', 'Actual Start Date', 'Submitted Date', 'Submitted By', 'Actual End Date', 'Completed By']
                Step2_df = pd.DataFrame(columns=columns)
                Temp=pd.read_excel(file_template_path) # Templates
                for index, _ in result_df.iterrows():
                    Temp_update=Temp.copy()
                    for i in range(len(Temp_update)):
                        Temp_update.at[i,'Task - Name']=result_df.at[index,'Task ID']
                        if i not in [1]:
                            Temp_update.at[i, 'Completed Date'] = result_df.loc[index, 'Initials / Date']
                            Temp_update.at[i, 'Completed By'] = result_df.loc[index, 'Submitted By SCTR']
                    Temp_update.at[0,'Parameter Reading 01']=result_df.at[index,'Rated Voltage']
                    Temp_update.at[0,'Parameter Reading 02']=result_df.at[index,'Ambient Temperature (℃)']
                    Temp_update.at[1,'Parameter Reading 01']=result_df.at[index,' Size (mm2)']
                    Temp_update.at[1,'Parameter Reading 02']=result_df.at[index,'Humidity (%)']
                    Temp_update.at[2,'Parameter Reading 01']=result_df.at[index,'Length (m)']
                    Temp_update.at[2,'Parameter Reading 02']=result_df.at[index,'Shield or Unshielded']
                    Temp_update.at[3,'Parameter Reading 01']=result_df.at[index,'Equipment']
                    Temp_update.at[3,'Parameter Reading 06']=result_df.at[index,' Make',]
                    Temp_update.at[3,'Parameter Reading 02']=result_df.at[index,'Model']
                    Temp_update.at[3,'Parameter Reading 03']=result_df.at[index,'Serial No.']
                    Temp_update.at[3,'Parameter Reading 04']=result_df.at[index,'Cal Expiry']
                    Temp_update.at[5,'Parameter Reading 01']=result_df.at[index,'U - V; W; E; Armour\n(MΩ)']
                    Temp_update.at[5,'Parameter Reading 02']=result_df.at[index,'U - V; W; E; Armour\nContinuity Test']
                    Temp_update.at[5,'Step Answer']=result_df.at[index,'U - V; W; E; Armour \n(Acc / Rej / NA)']
                    Temp_update.at[6,'Parameter Reading 01']=result_df.at[index,'V-U; W; E; Armour\n(MΩ)']
                    Temp_update.at[6,'Parameter Reading 02']=result_df.at[index,'V-U; W; E; Armour\nContinuity Test']
                    Temp_update.at[6,'Step Answer']=result_df.at[index,'V-U; W; E; Armour \n(Acc / Rej / NA)']
                    Temp_update.at[7,'Parameter Reading 01']=result_df.at[index,'W - U; V; E; Armour\n(MΩ)']
                    Temp_update.at[7,'Parameter Reading 02']=result_df.at[index,'W - U; V; E; Armour\nContinuity Test']
                    Temp_update.at[7,'Step Answer']=result_df.at[index,'W - U; V; E; Armour \n(Acc / Rej / NA)']
                    Temp_update.at[8,'Parameter Reading 01']=result_df.at[index,'N – U; V; W; E; Armour\n(MΩ) (If applicable)']
                    Temp_update.at[8,'Parameter Reading 02']=result_df.at[index,'N – U; V; W; E; Armour\nContinuity Test']
                    Temp_update.at[8,'Step Answer']=result_df.at[index,'N – U; V; W; E; Armour \n(Acc / Rej / NA)']
                    Temp_update.at[9,'Parameter Reading 01']=result_df.at[index,'Shield- Armour / Earth \n(MΩ)']
                    Temp_update.at[9,'Parameter Reading 02']=result_df.at[index,'Shield- Armour / Earth \nContinuity Test']
                    Temp_update.at[9,'Step Answer']=result_df.at[index,'Shield- Armour / Earth  \n(Acc / Rej / NA)']
                    Temp_update.at[11,'Parameter Reading 01']=result_df.at[index,'L- N /Armour / E\n(MΩ)']
                    Temp_update.at[11,'Parameter Reading 02']=result_df.at[index,'L- N /Armour / E\nContinuity Test']
                    Temp_update.at[11,'Step Answer']=result_df.at[index,'L- N /Armour / E\n(Acc / Rej / NA)']
                    Temp_update.at[12,'Parameter Reading 01']=result_df.at[index,'N-L / Armour / E\n (MΩ)']
                    Temp_update.at[12,'Parameter Reading 02']=result_df.at[index,'N-L / Armour / E\nContinuity Test']
                    Temp_update.at[12,'Step Answer']=result_df.at[index,'N-L / Armour / E\n(Acc / Rej / NA)']
                    Temp_update.at[14,'Parameter Reading 01']=result_df.at[index,'Core to Core (MΩ)']
                    Temp_update.at[14,'Parameter Reading 02']=result_df.at[index,'Core to Core\nContinuity Test']
                    Temp_update.at[14,'Step Answer']=result_df.at[index,'Core to Core\n(Acc / Rej / NA)']
                    Temp_update.at[15,'Parameter Reading 01']=result_df.at[index,'Core To Armour/Earth\n(MΩ)']
                    Temp_update.at[15,'Parameter Reading 02']=result_df.at[index,'Core To Armour/Earth\nContinuity Test']
                    Temp_update.at[15,'Step Answer']=result_df.at[index,'Core To Armour/Earth\n(Acc / Rej / NA)']
                    Temp_update.at[16,'Parameter Reading 01']=result_df.at[index,'Shield- Armour / Earth\n(MΩ)']
                    Temp_update.at[16,'Parameter Reading 02']=result_df.at[index,'Shield- Armour / Earth\nContinuity Test']
                    Temp_update.at[16,'Step Answer']=result_df.at[index,'Shield- Armour / Earth\n(Acc / Rej / NA)']
                    Temp_update['Step Sequence'] = Temp_update['Step Sequence'].astype(str)
                    Temp_update['Step Point'] = Temp_update['Step Point'].astype(str)
                    Temp_update['Step Action'] = Temp_update['Step Action'].astype(str)
                    Temp_update['Inspection Type'] = Temp_update['Inspection Type'].astype(str)
                    Step_1 = pd.concat([Step_1, Temp_update.copy()], ignore_index=True)
                    Step2_L=[None]*7
                    Step2_L[0]=result_df.loc[index,'Task ID']
                    Step2_L[1]=result_df.loc[index,'Task Model (Name)']
                    Step2_L[2]=result_df.loc[index,'Initials / Date']
                    Step2_L[3]=result_df.loc[index,'Submitted Date']
                    Step2_L[4]=result_df.loc[index,'Submitted By SCTR']
                    Step2_L[5]=result_df.loc[index,'Completed Date']
                    Step2_L[6]=result_df.loc[index,'Completed By CTR']
                    # Create the new row to a DataFrame
                    new_row_df = pd.DataFrame([Step2_L], columns=columns)
                    Step2_df = pd.concat([Step2_df, new_row_df], ignore_index=True)

                # Step 1: Replace 'nan' with empty string in the "Step Action" column
                Step_1["Step Action"] = Step_1["Step Action"].str.replace('nan', '')

                # Step 2: Write Step_1 DataFrame to an Excel file in memory and apply formatting
                excel_file2 = io.BytesIO()
                with pd.ExcelWriter(excel_file2, engine='openpyxl') as writer:
                    Step_1.to_excel(writer, index=False, sheet_name='Sheet1')
                    worksheet = writer.sheets['Sheet1']
                                            
                    # Apply the custom date format for specific columns
                    date_format = 'DD-MMM-YY'
                    date_columns = ['Completed Date','Parameter Reading 04']
                    for col in date_columns:
                        col_index = Step_1.columns.get_loc(col)
                        for row in range(2, worksheet.max_row + 1):  # Assuming header is in the first row
                            worksheet.cell(row=row, column=col_index + 1).number_format = date_format

                    # Define the column names
                    column_names = ['Completed By', 'Completed Date']

                    # Get the column indices from the column names
                    column_indices = {}
                    for column_name in column_names:
                        for col in worksheet.iter_cols(min_row=1, max_row=1):
                            for cell in col:
                                if cell.value == column_name:
                                    column_indices[column_name] = cell.col_idx  # Use col_idx to get the column index
                                    break

                    # Skip if any column not found
                    for column_name in column_names:
                        if column_name not in column_indices:
                            print(f"Column '{column_name}' not found in the worksheet.")
                            continue

                    # Iterate over each row
                    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                        completed_date_cell = worksheet.cell(row=row_num, column=column_indices['Completed Date'])  # Get the 'Completed Date' cell

                        # Check if 'Completed Date' cell is None
                        if completed_date_cell.value is None:
                            continue  # Skip the row if 'Completed Date' cell is None

                        completed_by_cell = worksheet.cell(row=row_num, column=column_indices['Completed By'])  # Get the 'Completed By' cell

                        # Check if 'Completed By' cell is None
                        if completed_by_cell.value is None:
                            completed_by_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

                    # Define the column to apply conditional formatting
                    column_name = 'Step Answer'

                    # Get the column index from the column name
                    column_index = None
                    for col in worksheet.iter_cols(min_row=1, max_row=1):
                        for cell in col:
                            if cell.value == column_name:
                                column_index = cell.col_idx  # Use col_idx to get the column index
                                break
                        if column_index:
                            break

                    # If column index is found, iterate through cells in the specified column to apply conditional formatting
                    if column_index:
                        for row in worksheet.iter_rows(min_row=2, min_col=column_index, max_row=worksheet.max_row, max_col=column_index):
                            for cell in row:
                                if cell.value not in ['Accept', 'Reject', 'N/A',None] and cell.value:
                                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    else:
                        print(f"Column '{column_name}' not found in the worksheet.")

                    # Move the file pointer to the beginning of the in-memory file
                    excel_file2.seek(0)
                
                # Read the mapping file into a DataFrame
                mapping_file = pd.read_excel(file_scdb_mapping_path)

                # Create a preprocessed dictionary with case-insensitive keys and removed white spaces
                preprocessed_dictionary = dict(
                    zip(
                        map(lambda x: "".join(str(x).lower().split()), mapping_file.iloc[:, 0]),
                        mapping_file.iloc[:, 1]
                    )
                )

                # Process Step2_df DataFrame
                Step2_df['Completed By'] = (
                    Step2_df['Completed By']
                    .map(lambda x: "".join(str(x).lower().split()))
                    .map(preprocessed_dictionary)
                )
                Step2_df = Step2_df.astype(str)
                Step2_df["Completed By"] = Step2_df["Completed By"].str.replace('nan', '')

                # Create an in-memory Excel file for Step2_df
                excel_file3 = io.BytesIO()
                with pd.ExcelWriter(excel_file3, engine='openpyxl') as writer:
                    Step2_df.to_excel(writer, index=False, sheet_name='Sheet1')
                    apply_formatting_missing(writer, 'Sheet1', Step2_df, ["Completed By"], "FF0000")
                excel_file3.seek(0)
                st.success("Files generated successfully!")
                st.write("Download your files here:")
                zip_file = io.BytesIO()
                with zipfile.ZipFile(zip_file, 'w') as zipf:
                    zipf.writestr("Report.xlsx", excel_file1.getvalue())
                    zipf.writestr("Step-1.xlsx", excel_file2.getvalue())
                    zipf.writestr("Step-2.xlsx", excel_file3.getvalue())

                # Provide a download link for the zip file
                st.download_button(
                    label="Download All Files",
                    data=zip_file,
                    file_name='all_files_309.zip',
                    mime='application/zip'
                )
                
            except Exception as e:
                st.error(f"Error executing logic and generating Excel files: {e}")
        
        else:
            st.warning("Please upload all files before executing logic.")
